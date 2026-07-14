# -*- coding: utf-8 -*-
"""
Appends scraped Career Fair rows into a student's tracker spreadsheet.

Usage (called by Claude, not run standalone with hardcoded data):
    python update_tracker.py <path_to_tracker.xlsx> <path_to_rows.json>

<path_to_rows.json> is a JSON array of objects, one per row, using these keys
(all required, use "" for unknown values, never omit a key):
    platform, employer, industry, job_position, job_description,
    job_type, location, opt_cpt, relevance_note

Matches existing rows by "employer" name (case-insensitive) and updates them
in place rather than duplicating; new employers are appended at the bottom.
New rows copy the font/alignment from the row directly above them, so a
student's manual formatting is never overwritten.
"""
import sys
import json
from openpyxl import load_workbook
from copy import copy

COLUMNS = [
    "platform", "employer", "industry", "job_position", "job_description",
    "job_type", "location", "opt_cpt", "relevance_note",
]


def load_rows(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    if len(sys.argv) != 3:
        print("Usage: python update_tracker.py <tracker.xlsx> <rows.json>")
        sys.exit(1)

    tracker_path, rows_path = sys.argv[1], sys.argv[2]
    rows = load_rows(rows_path)

    wb = load_workbook(tracker_path)
    ws = wb.active

    header = [c.value for c in ws[1]]
    if [h.strip().lower() if h else "" for h in header[:2]] != ["platform", "employer"]:
        print("Header row doesn't look like the expected template (Platform, Employer, ...).")
        print("Found header:", header)
        print("Fix the header or point this script at the correct sheet before continuing.")
        sys.exit(1)

    existing_by_employer = {}
    for r in range(2, ws.max_row + 1):
        employer_cell = ws.cell(row=r, column=2).value
        if employer_cell:
            existing_by_employer[employer_cell.strip().lower()] = r

    last_row = ws.max_row
    added, updated = 0, 0

    for row_data in rows:
        employer_key = row_data.get("employer", "").strip().lower()
        if not employer_key:
            continue

        if employer_key in existing_by_employer:
            target_row = existing_by_employer[employer_key]
            updated += 1
        else:
            last_row += 1
            target_row = last_row
            added += 1
            ref_row = target_row - 1 if target_row > 2 else 2
            for col_idx in range(1, len(COLUMNS) + 1):
                ref_cell = ws.cell(row=ref_row, column=col_idx)
                new_cell = ws.cell(row=target_row, column=col_idx)
                new_cell.font = copy(ref_cell.font)
                new_cell.alignment = copy(ref_cell.alignment)

        for col_idx, key in enumerate(COLUMNS, start=1):
            ws.cell(row=target_row, column=col_idx, value=row_data.get(key, ""))

    wb.save(tracker_path)
    print(f"Done. Added {added} new employer(s), updated {updated} existing row(s).")


if __name__ == "__main__":
    main()
